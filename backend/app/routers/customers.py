from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from pydantic import BaseModel
from datetime import datetime
from app.core.database import get_db
from app.core.deps import get_current_user, check_permission, check_permission, log_activity
from app.models.customer import Customer
from app.models.order import Order
from app.models.user import User

router = APIRouter(prefix="/customers", tags=["customers"])

class CustomerCreate(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    payment_term_days: Optional[int] = None

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    payment_term_days: Optional[int] = None

def build_customer_out(c: Customer, db: Session) -> dict:
    order_count = db.query(func.count(Order.id)).filter(Order.customer_id == c.id).scalar()
    total_spent = db.query(func.sum(Order.id)).filter(Order.customer_id == c.id).scalar()
    return {
        "id": c.id,
        "name": c.name,
        "phone": c.phone,
        "email": c.email,
        "address": c.address,
        "notes": c.notes,
        "payment_term_days": c.payment_term_days,
        "created_at": c.created_at,
        "order_count": order_count or 0,
    }

@router.get("/export/prices")
def export_all_prices_route(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from app.models.customer_price import CustomerPrice
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    prices = db.query(CustomerPrice).order_by(
        CustomerPrice.customer_id, CustomerPrice.product_id
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ozel Fiyatlar"

    hf = PatternFill("solid", fgColor="1e40af")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    alt = PatternFill("solid", fgColor="f0f4ff")

    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value="Laves Kimya - Musteri Ozel Fiyat Listesi")
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = hf
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    headers = ["Musteri", "Urun", "Standart Fiyat (TL)", "Ozel Fiyat (TL)", "Fark (TL)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    for i, p in enumerate(prices, 3):
        row_fill = alt if i % 2 == 0 else None
        std = p.product.sale_price if p.product else 0
        diff = p.price - std
        vals = [
            p.customer.name if p.customer else "-",
            p.product.name if p.product else "-",
            std, p.price, round(diff, 2)
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            if row_fill: cell.fill = row_fill
            if col in (3, 4, 5): cell.number_format = '#,##0.00'
            if col == 5 and diff < 0:
                cell.font = Font(color="dc2626")
            elif col == 5 and diff > 0:
                cell.font = Font(color="16a34a")

    col_widths = [24, 24, 20, 20, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ozel_fiyatlar.xlsx"})


@router.get("")
def list_customers(
    skip: int = 0, limit: int = 20,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_permission("customers", "view"))
):
    query = db.query(Customer)
    if search:
        query = query.filter(Customer.name.ilike(f"%{search}%"))
    total = query.count()
    items = query.order_by(Customer.name).offset(skip).limit(limit).all()
    return {"total": total, "items": [build_customer_out(c, db) for c in items]}

@router.post("")
def create_customer(data: CustomerCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    customer = Customer(**data.model_dump())
    db.add(customer)
    db.commit()
    db.refresh(customer)
    log_activity(db, current_user.id, "CREATE", "Customer", customer.id, f"Müşteri oluşturuldu: {customer.name}")
    return build_customer_out(customer, db)

@router.get("/{customer_id}")
def get_customer(customer_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    orders = db.query(Order).filter(Order.customer_id == customer_id).order_by(Order.created_at.desc()).all()
    result = build_customer_out(c, db)
    result["orders"] = [{"id": o.id, "status": o.status, "created_at": o.created_at, "notes": o.notes} for o in orders]
    return result

@router.put("/{customer_id}")
def update_customer(customer_id: int, data: CustomerUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    for key, value in data.model_dump(exclude_none=True).items():
        setattr(c, key, value)
    db.commit()
    db.refresh(c)
    log_activity(db, current_user.id, "UPDATE", "Customer", customer_id)
    return build_customer_out(c, db)

@router.delete("/{customer_id}")
def delete_customer(customer_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Müşteri bulunamadı")
    db.delete(c)
    db.commit()
    log_activity(db, current_user.id, "DELETE", "Customer", customer_id)
    return {"message": "Müşteri silindi"}


# ── Müşteri Özel Fiyatları ─────────────────────────────────────────────────

class CustomerPriceSet(BaseModel):
    product_id: int
    price: float

@router.get("/{customer_id}/prices")
def get_customer_prices(customer_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from app.models.customer_price import CustomerPrice
    from app.models.product import Product
    prices = db.query(CustomerPrice).filter(CustomerPrice.customer_id == customer_id).all()
    return [{"id": p.id, "product_id": p.product_id,
             "product_name": p.product.name if p.product else None,
             "standard_price": p.product.sale_price if p.product else None,
             "price": p.price} for p in prices]

@router.put("/{customer_id}/prices")
def set_customer_prices(customer_id: int, prices: list[CustomerPriceSet],
                        db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from app.models.customer_price import CustomerPrice
    c = db.query(Customer).filter(Customer.id == customer_id).first()
    if not c: raise HTTPException(404, "Müşteri bulunamadı")
    # Mevcut fiyatları sil, yeniden yaz
    db.query(CustomerPrice).filter(CustomerPrice.customer_id == customer_id).delete()
    for p in prices:
        db.add(CustomerPrice(customer_id=customer_id, product_id=p.product_id, price=p.price))
    db.commit()
    return {"message": f"{len(prices)} fiyat kaydedildi"}

@router.get("/{customer_id}/price-for/{product_id}")
def get_price_for_product(customer_id: int, product_id: int,
                          db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Sipariş oluştururken kullanılır — özel fiyat varsa onu döner"""
    from app.models.customer_price import CustomerPrice
    from app.models.product import Product
    cp = db.query(CustomerPrice).filter(
        CustomerPrice.customer_id == customer_id,
        CustomerPrice.product_id == product_id
    ).first()
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product: raise HTTPException(404, "Ürün bulunamadı")
    return {
        "product_id": product_id,
        "standard_price": product.sale_price,
        "custom_price": cp.price if cp else None,
        "effective_price": cp.price if cp else product.sale_price
    }



