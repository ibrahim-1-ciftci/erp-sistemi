from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from datetime import date
from io import BytesIO
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.database import get_db
from app.core.deps import get_current_user, log_activity
from app.models.debt import Debt, DebtStatus
from app.models.user import User

router = APIRouter(prefix="/debts", tags=["debts"])

class DebtCreate(BaseModel):
    creditor: str
    description: Optional[str] = None
    total_amount: float
    due_date: date
    notes: Optional[str] = None

class DebtUpdate(BaseModel):
    creditor: Optional[str] = None
    description: Optional[str] = None
    total_amount: Optional[float] = None
    paid_amount: Optional[float] = None
    due_date: Optional[date] = None
    paid_date: Optional[date] = None
    notes: Optional[str] = None

class PaymentIn(BaseModel):
    amount: float
    paid_date: Optional[date] = None

def compute_status(d: Debt) -> str:
    paid = d.paid_amount or 0
    if paid >= d.total_amount:
        return DebtStatus.paid
    if paid > 0:
        return DebtStatus.partial
    if d.due_date < date.today():
        return DebtStatus.overdue
    return DebtStatus.pending

def build(d: Debt) -> dict:
    status = compute_status(d)
    remaining = d.total_amount - (d.paid_amount or 0)
    overdue_days = (date.today() - d.due_date).days if d.due_date < date.today() and status != DebtStatus.paid else 0
    return {
        "id": d.id, "creditor": d.creditor, "description": d.description,
        "total_amount": d.total_amount, "paid_amount": d.paid_amount or 0,
        "remaining": remaining, "due_date": d.due_date, "paid_date": d.paid_date,
        "status": status, "notes": d.notes, "created_at": d.created_at,
        "overdue_days": overdue_days,
    }

@router.get("")
def list_debts(status: Optional[str] = None, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    debts = db.query(Debt).order_by(Debt.due_date).all()
    result = [build(d) for d in debts]
    if status:
        result = [r for r in result if r["status"] == status]
    total_debt     = sum(r["remaining"] for r in result if r["status"] != DebtStatus.paid)
    overdue_count  = sum(1 for r in result if r["status"] == DebtStatus.overdue)
    return {"items": result, "total_debt": total_debt, "overdue_count": overdue_count}

@router.post("")
def create_debt(data: DebtCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    d = Debt(**data.model_dump())
    db.add(d); db.commit(); db.refresh(d)
    log_activity(db, current_user.id, "CREATE", "Debt", d.id, f"Borç: {d.creditor}")
    return build(d)

@router.put("/{debt_id}")
def update_debt(debt_id: int, data: DebtUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    d = db.query(Debt).filter(Debt.id == debt_id).first()
    if not d: raise HTTPException(404, "Bulunamadı")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(d, k, v)
    db.commit(); db.refresh(d)
    return build(d)

@router.post("/{debt_id}/pay")
def pay_debt(debt_id: int, data: PaymentIn, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    d = db.query(Debt).filter(Debt.id == debt_id).first()
    if not d: raise HTTPException(404, "Bulunamadı")
    d.paid_amount = (d.paid_amount or 0) + data.amount
    if d.paid_amount >= d.total_amount and not d.paid_date:
        d.paid_date = data.paid_date or date.today()
    db.commit(); db.refresh(d)
    log_activity(db, current_user.id, "PAY", "Debt", debt_id, f"Ödeme: {data.amount}")
    return build(d)

@router.delete("/{debt_id}")
def delete_debt(debt_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    d = db.query(Debt).filter(Debt.id == debt_id).first()
    if not d: raise HTTPException(404, "Bulunamadı")
    db.delete(d); db.commit()
    return {"message": "Silindi"}

@router.get("/export")
def export_debts(status: Optional[str] = None, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    debts = db.query(Debt).order_by(Debt.due_date).all()
    result = [build(d) for d in debts]
    if status:
        result = [r for r in result if r["status"] == status]

    STATUS_TR = {"pending": "Ödenmedi", "partial": "Kısmi", "paid": "Ödendi", "overdue": "Gecikmiş"}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Borç Takibi"

    hf = PatternFill("solid", fgColor="1e40af")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    alt = PatternFill("solid", fgColor="f0f4ff")
    red = PatternFill("solid", fgColor="fee2e2")

    # Şirket başlığı
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value="Laves Kimya - Borç Takibi")
    c.font = Font(bold=True, size=13, color="FFFFFF"); c.fill = hf
    c.alignment = Alignment(horizontal="center"); ws.row_dimensions[1].height = 22

    headers = ["Alacaklı", "Açıklama", "Toplam (₺)", "Ödenen (₺)", "Kalan (₺)", "Vade Tarihi", "Durum", "Notlar"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(result, 3):
        row_fill = red if r["status"] == "overdue" else (alt if i % 2 == 0 else None)
        vals = [r["creditor"], r["description"] or "", r["total_amount"],
                r["paid_amount"], r["remaining"],
                str(r["due_date"]), STATUS_TR.get(r["status"], r["status"]),
                r["notes"] or ""]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            if row_fill: cell.fill = row_fill
            if col in (3, 4, 5): cell.number_format = '#,##0.00'

    col_widths = [22, 28, 14, 14, 14, 14, 12, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=borc_takibi.xlsx"})
