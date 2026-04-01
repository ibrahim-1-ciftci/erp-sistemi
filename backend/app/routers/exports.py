from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.raw_material import RawMaterial
from app.models.product import Product
from app.models.order import Order
from app.models.customer import Customer
from app.models.supplier import Supplier
from app.models.user import User

router = APIRouter(prefix="/exports", tags=["exports"])

COMPANY = "Laves Kimya"
H_FILL  = PatternFill("solid", fgColor="1e40af")
H_FONT  = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL= PatternFill("solid", fgColor="f0f4ff")

def make_wb(sheet_title: str, headers: list) -> tuple:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    # Şirket başlığı
    ws.merge_cells(f"A1:{openpyxl.utils.get_column_letter(len(headers))}1")
    c = ws.cell(row=1, column=1, value=COMPANY)
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = H_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    # Başlık satırı
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center")
    return wb, ws

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except: pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

def stream(wb, filename: str):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/raw-materials")
def export_raw_materials(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    headers = ["ID", "Ad", "Birim", "Stok Miktarı", "Min Stok", "Alış Fiyatı (₺)", "Tedarikçi", "Kayıt Tarihi"]
    wb, ws = make_wb("Hammaddeler", headers)
    items = db.query(RawMaterial).order_by(RawMaterial.name).all()
    for i, m in enumerate(items, 3):
        fill = ALT_FILL if i % 2 == 0 else None
        vals = [m.id, m.name, m.unit, m.stock_quantity, m.min_stock_level,
                m.purchase_price, m.supplier.name if m.supplier else "-",
                m.created_at.strftime("%d.%m.%Y") if m.created_at else ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            if fill: c.fill = fill
    auto_width(ws)
    return stream(wb, "hammaddeler.xlsx")


@router.get("/products")
def export_products(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    headers = ["ID", "Ürün Adı", "Satış Fiyatı (₺)", "Stok Miktarı", "Kayıt Tarihi"]
    wb, ws = make_wb("Ürünler", headers)
    items = db.query(Product).order_by(Product.name).all()
    for i, p in enumerate(items, 3):
        fill = ALT_FILL if i % 2 == 0 else None
        vals = [p.id, p.name, p.sale_price, p.stock_quantity,
                p.created_at.strftime("%d.%m.%Y") if p.created_at else ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            if fill: c.fill = fill
    auto_width(ws)
    return stream(wb, "urunler.xlsx")


@router.get("/orders")
def export_orders(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    headers = ["Sipariş No", "Müşteri", "Telefon", "E-posta", "Ürünler", "Toplam (₺)", "Durum", "Tarih"]
    wb, ws = make_wb("Siparişler", headers)
    STATUS_TR = {"pending": "Bekliyor", "in_production": "Üretimde",
                 "completed": "Tamamlandı", "shipped": "Sevkiyatta", "cancelled": "İptal"}
    orders = db.query(Order).order_by(Order.created_at.desc()).all()
    for i, o in enumerate(orders, 3):
        fill = ALT_FILL if i % 2 == 0 else None
        total = sum((item.unit_price or 0) * item.quantity for item in o.items)
        products_str = ", ".join(item.product.name for item in o.items if item.product)
        vals = [f"#{o.id}", o.customer_name, o.customer_phone or "", o.customer_email or "",
                products_str, round(total, 2), STATUS_TR.get(str(o.status), str(o.status)),
                o.created_at.strftime("%d.%m.%Y") if o.created_at else ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            if fill: c.fill = fill
    auto_width(ws)
    return stream(wb, "siparisler.xlsx")


@router.get("/customers")
def export_customers(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    headers = ["ID", "Müşteri Adı", "Telefon", "E-posta", "Adres", "Vade Günü", "Sipariş Sayısı", "Kayıt Tarihi"]
    wb, ws = make_wb("Müşteriler", headers)
    from sqlalchemy import func
    customers = db.query(Customer).order_by(Customer.name).all()
    for i, c in enumerate(customers, 3):
        fill = ALT_FILL if i % 2 == 0 else None
        order_count = db.query(func.count(Order.id)).filter(Order.customer_id == c.id).scalar() or 0
        vals = [c.id, c.name, c.phone or "", c.email or "", c.address or "",
                f"{c.payment_term_days} gün" if c.payment_term_days else "-",
                order_count, c.created_at.strftime("%d.%m.%Y") if c.created_at else ""]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            if fill: cell.fill = fill
    auto_width(ws)
    return stream(wb, "musteriler.xlsx")


@router.get("/suppliers")
def export_suppliers(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    headers = ["ID", "Tedarikçi Adı", "Telefon", "E-posta", "Adres", "Kayıt Tarihi"]
    wb, ws = make_wb("Tedarikçiler", headers)
    suppliers = db.query(Supplier).order_by(Supplier.name).all()
    for i, s in enumerate(suppliers, 3):
        fill = ALT_FILL if i % 2 == 0 else None
        vals = [s.id, s.name, s.phone or "", s.email or "", s.address or "",
                s.created_at.strftime("%d.%m.%Y") if s.created_at else ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            if fill: c.fill = fill
    auto_width(ws)
    return stream(wb, "tedarikciler.xlsx")
