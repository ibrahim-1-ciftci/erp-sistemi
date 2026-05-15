from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from typing import Optional
from datetime import date, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.core.database import get_db
from app.core.deps import get_current_user, check_permission
from app.models.raw_material import RawMaterial
from app.models.product import Product
from app.models.order import Order, OrderStatus
from app.models.production import Production, ProductionStatus
from app.models.stock_movement import StockMovement
from app.models.activity_log import ActivityLog
from app.models.user import User

router = APIRouter(prefix="/reports", tags=["reports"])

COMPANY_NAME = "Laves Kimya"

def add_company_header(ws, col_count: int):
    ws.merge_cells(f"A1:{openpyxl.utils.get_column_letter(col_count)}1")
    c = ws.cell(row=1, column=1, value=COMPANY_NAME)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1e40af")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    return 3  # veri başlangıç satırı


@router.get("/dashboard")
def dashboard(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from app.models.payment import Payment
    from app.models.debt import Debt
    from app.models.customer import Customer
    from datetime import date as dt_date

    # Siparişler
    pending_orders    = db.query(Order).filter(Order.status == OrderStatus.pending).count()
    completed_orders  = db.query(Order).filter(Order.status == OrderStatus.completed).count()
    shipped_orders    = db.query(Order).filter(Order.status == OrderStatus.shipped).count()
    total_orders      = db.query(Order).count()

    # Bu ay siparişler
    month_start = dt_date.today().replace(day=1)
    month_orders = db.query(Order).filter(func.date(Order.created_at) >= month_start).count()

    # Ürünler
    products = db.query(Product).all()
    total_products = len(products)

    # Müşteriler
    total_customers = db.query(Customer).count()

    # Vade (alacak)
    payments = db.query(Payment).all()
    total_receivable = sum(p.total_amount - (p.paid_amount or 0) for p in payments if (p.paid_amount or 0) < p.total_amount)
    overdue_payments = sum(1 for p in payments if p.due_date < dt_date.today() and (p.paid_amount or 0) < p.total_amount)

    # Borç
    debts = db.query(Debt).all()
    total_payable = sum(d.total_amount - (d.paid_amount or 0) for d in debts if (d.paid_amount or 0) < d.total_amount)
    overdue_debts = sum(1 for d in debts if d.due_date < dt_date.today() and (d.paid_amount or 0) < d.total_amount)

    # Son 6 ay sipariş sayısı (grafik)
    monthly_orders = []
    for i in range(5, -1, -1):
        d = dt_date.today().replace(day=1) - timedelta(days=i*28)
        ms = d.replace(day=1)
        if ms.month == 12:
            me = ms.replace(year=ms.year+1, month=1, day=1)
        else:
            me = ms.replace(month=ms.month+1, day=1)
        count = db.query(Order).filter(
            func.date(Order.created_at) >= ms,
            func.date(Order.created_at) < me
        ).count()
        # Bu aydaki toplam sipariş tutarı
        orders_in_month = db.query(Order).filter(
            func.date(Order.created_at) >= ms,
            func.date(Order.created_at) < me
        ).all()
        monthly_orders.append({
            "month": ms.strftime("%b %Y"),
            "count": count,
        })

    # Son 5 sipariş
    recent_orders = db.query(Order).order_by(Order.created_at.desc()).limit(5).all()
    recent_orders_out = []
    for o in recent_orders:
        total_val = sum((i.unit_price or 0) * i.quantity for i in o.items)
        recent_orders_out.append({
            "id": o.id,
            "customer_name": o.customer_name,
            "status": o.status,
            "total_value": round(total_val, 2),
            "created_at": o.created_at,
        })

    return {
        "pending_orders":    pending_orders,
        "completed_orders":  completed_orders,
        "shipped_orders":    shipped_orders,
        "total_orders":      total_orders,
        "month_orders":      month_orders,
        "total_products":    total_products,
        "total_customers":   total_customers,
        "total_receivable":  round(total_receivable, 2),
        "overdue_payments":  overdue_payments,
        "total_payable":     round(total_payable, 2),
        "overdue_debts":     overdue_debts,
        "monthly_orders":    monthly_orders,
        "recent_orders":     recent_orders_out,
    }

@router.get("/stock-movements")
def stock_movements(
    skip: int = 0, limit: int = 100,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    search: Optional[str] = None,
    movement_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(StockMovement)
    if date_from:
        query = query.filter(func.date(StockMovement.created_at) >= date_from)
    if date_to:
        query = query.filter(func.date(StockMovement.created_at) <= date_to)
    if movement_type:
        query = query.filter(StockMovement.type == movement_type)
    total = query.count()
    movements = query.order_by(StockMovement.created_at.desc()).offset(skip).limit(limit).all()
    result = []
    for mv in movements:
        mat_name = mv.material.name if mv.material else None
        prd_name = mv.product.name if mv.product else None
        name = mat_name or prd_name or ''
        # Arama filtresi — Python tarafında (join yazmaktan kaçınmak için)
        if search and search.lower() not in name.lower() and search.lower() not in (mv.description or '').lower():
            continue
        result.append({
            "id": mv.id,
            "type": mv.type,
            "quantity": mv.quantity,
            "description": mv.description,
            "created_at": mv.created_at,
            "material_name": mat_name,
            "product_name": prd_name,
        })
    return {"total": total, "items": result}

@router.get("/activity-logs")
def activity_logs(
    skip: int = 0, limit: int = 100,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    search: Optional[str] = None,
    action_filter: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(ActivityLog)
    if date_from:
        query = query.filter(func.date(ActivityLog.created_at) >= date_from)
    if date_to:
        query = query.filter(func.date(ActivityLog.created_at) <= date_to)
    if action_filter:
        query = query.filter(ActivityLog.action == action_filter)
    if search:
        query = query.filter(
            ActivityLog.details.ilike(f"%{search}%") |
            ActivityLog.entity.ilike(f"%{search}%") |
            ActivityLog.action.ilike(f"%{search}%")
        )
    total = query.count()
    logs = query.order_by(ActivityLog.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": logs}


@router.delete("/cleanup")
def cleanup_old_logs(
    days: int = 15,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """15 günden eski activity_log ve stock_movement kayıtlarını sil."""
    from datetime import datetime, timezone
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    deleted_logs = db.query(ActivityLog).filter(ActivityLog.created_at < cutoff).delete()
    deleted_movements = db.query(StockMovement).filter(StockMovement.created_at < cutoff).delete()
    db.commit()
    return {"deleted_logs": deleted_logs, "deleted_movements": deleted_movements, "cutoff_days": days}

@router.get("/profitability")
def profitability(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Production).filter(Production.status == ProductionStatus.completed)
    if date_from:
        query = query.filter(func.date(Production.completed_at) >= date_from)
    if date_to:
        query = query.filter(func.date(Production.completed_at) <= date_to)
    productions = query.all()
    result = []
    for p in productions:
        revenue = p.product.sale_price * p.quantity if p.product else 0
        profit = revenue - p.total_cost
        result.append({
            "production_id": p.id,
            "product_name": p.product.name if p.product else None,
            "quantity": p.quantity,
            "total_cost": p.total_cost,
            "revenue": revenue,
            "profit": profit,
            "margin_pct": (profit / revenue * 100) if revenue > 0 else 0,
            "completed_at": p.completed_at
        })
    return result

@router.get("/export/stock-movements")
def export_stock_movements(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(StockMovement)
    if date_from:
        query = query.filter(func.date(StockMovement.created_at) >= date_from)
    if date_to:
        query = query.filter(func.date(StockMovement.created_at) <= date_to)
    movements = query.order_by(StockMovement.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stok Hareketleri"

    header_fill = PatternFill("solid", fgColor="1e40af")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Tarih", "Tür", "Malzeme/Ürün", "Miktar", "Açıklama"]
    start_row = add_company_header(ws, len(headers))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, mv in enumerate(movements, start_row + 1):
        ws.cell(row=row, column=1, value=mv.created_at.strftime("%d.%m.%Y %H:%M") if mv.created_at else "")
        ws.cell(row=row, column=2, value="Giriş" if mv.type == "in" else "Çıkış")
        ws.cell(row=row, column=3, value=mv.material.name if mv.material else (mv.product.name if mv.product else "-"))
        ws.cell(row=row, column=4, value=mv.quantity)
        ws.cell(row=row, column=5, value=mv.description or "")

    for col in ws.columns:
        first_cell = col[0]
        if hasattr(first_cell, 'column_letter'):
            ws.column_dimensions[first_cell.column_letter].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stok_hareketleri.xlsx"}
    )

@router.get("/export/profitability")
def export_profitability(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Production).filter(Production.status == ProductionStatus.completed)
    if date_from:
        query = query.filter(func.date(Production.completed_at) >= date_from)
    if date_to:
        query = query.filter(func.date(Production.completed_at) <= date_to)
    productions = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Karlılık Raporu"

    header_fill = PatternFill("solid", fgColor="1e40af")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Ürün", "Adet", "Maliyet (₺)", "Gelir (₺)", "Kâr (₺)", "Marj (%)", "Tarih"]
    start_row = add_company_header(ws, len(headers))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, p in enumerate(productions, start_row + 1):
        revenue = p.product.sale_price * p.quantity if p.product else 0
        profit = revenue - p.total_cost
        margin = (profit / revenue * 100) if revenue > 0 else 0
        ws.cell(row=row, column=1, value=p.product.name if p.product else "-")
        ws.cell(row=row, column=2, value=p.quantity)
        ws.cell(row=row, column=3, value=round(p.total_cost, 2))
        ws.cell(row=row, column=4, value=round(revenue, 2))
        ws.cell(row=row, column=5, value=round(profit, 2))
        ws.cell(row=row, column=6, value=round(margin, 1))
        ws.cell(row=row, column=7, value=p.completed_at.strftime("%d.%m.%Y") if p.completed_at else "")

    for col in ws.columns:
        first_cell = col[0]
        if hasattr(first_cell, 'column_letter'):
            ws.column_dimensions[first_cell.column_letter].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=karlilik_raporu.xlsx"}
    )
