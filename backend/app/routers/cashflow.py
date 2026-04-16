from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import date
from io import BytesIO
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.core.database import get_db
from app.core.deps import get_current_user, check_permission, check_permission, log_activity
from app.models.cashflow import CashFlow
from app.models.user import User

router = APIRouter(prefix="/cashflow", tags=["cashflow"])

class FlowCreate(BaseModel):
    flow_date:   date
    flow_type:   str          # income | expense
    pay_method:  Optional[str] = None
    amount:      float
    description: Optional[str] = None
    category:    Optional[str] = None

class FlowUpdate(BaseModel):
    flow_date:   Optional[date] = None
    flow_type:   Optional[str] = None
    pay_method:  Optional[str] = None
    amount:      Optional[float] = None
    description: Optional[str] = None
    category:    Optional[str] = None

def build(f: CashFlow) -> dict:
    return {
        "id": f.id, "flow_date": f.flow_date, "flow_type": f.flow_type,
        "pay_method": f.pay_method, "amount": f.amount,
        "description": f.description, "category": f.category,
        "created_at": f.created_at,
    }

@router.get("")
def list_flows(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    flow_type: Optional[str]  = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_permission("cashflow", "view"))
):
    query = db.query(CashFlow).order_by(CashFlow.flow_date.desc(), CashFlow.created_at.desc())
    if date_from: query = query.filter(CashFlow.flow_date >= date_from)
    if date_to:   query = query.filter(CashFlow.flow_date <= date_to)
    if flow_type: query = query.filter(CashFlow.flow_type == flow_type)
    items = query.all()

    total_income  = sum(f.amount for f in items if f.flow_type == "income")
    total_expense = sum(f.amount for f in items if f.flow_type == "expense")

    # Ödeme yöntemine göre gelir dağılımı
    cash_income  = sum(f.amount for f in items if f.flow_type == "income" and f.pay_method == "cash")
    pos_income   = sum(f.amount for f in items if f.flow_type == "income" and f.pay_method == "pos")
    elden_income = sum(f.amount for f in items if f.flow_type == "income" and f.pay_method == "elden")

    return {
        "items": [build(f) for f in items],
        "total_income": round(total_income, 2),
        "total_expense": round(total_expense, 2),
        "net": round(total_income - total_expense, 2),
        "cash_income": round(cash_income, 2),
        "pos_income": round(pos_income, 2),
        "elden_income": round(elden_income, 2),
    }

@router.get("/daily-summary")
def daily_summary(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Günlük özet — grafik için"""
    query = db.query(CashFlow)
    if date_from: query = query.filter(CashFlow.flow_date >= date_from)
    if date_to:   query = query.filter(CashFlow.flow_date <= date_to)
    items = query.all()

    by_day = {}
    for f in items:
        d = str(f.flow_date)
        if d not in by_day:
            by_day[d] = {"date": d, "income": 0, "expense": 0}
        if f.flow_type == "income":
            by_day[d]["income"] += f.amount
        else:
            by_day[d]["expense"] += f.amount

    result = sorted(by_day.values(), key=lambda x: x["date"])
    for r in result:
        r["net"] = round(r["income"] - r["expense"], 2)
        r["income"] = round(r["income"], 2)
        r["expense"] = round(r["expense"], 2)
    return result

@router.post("")
def create_flow(data: FlowCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    f = CashFlow(**data.model_dump())
    db.add(f); db.commit(); db.refresh(f)
    log_activity(db, current_user.id, "CREATE", "CashFlow", f.id,
                 f"{'Gelir' if data.flow_type=='income' else 'Gider'}: ₺{data.amount}")
    return build(f)

@router.put("/{flow_id}")
def update_flow(flow_id: int, data: FlowUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    f = db.query(CashFlow).filter(CashFlow.id == flow_id).first()
    if not f: raise HTTPException(404, "Kayıt bulunamadı")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(f, k, v)
    db.commit(); db.refresh(f)
    return build(f)

@router.delete("/{flow_id}")
def delete_flow(flow_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    f = db.query(CashFlow).filter(CashFlow.id == flow_id).first()
    if not f: raise HTTPException(404, "Kayıt bulunamadı")
    db.delete(f); db.commit()
    return {"message": "Silindi"}

@router.get("/export")
def export_cashflow(
    date_from: Optional[date] = None,
    date_to:   Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(CashFlow).order_by(CashFlow.flow_date.desc())
    if date_from: query = query.filter(CashFlow.flow_date >= date_from)
    if date_to:   query = query.filter(CashFlow.flow_date <= date_to)
    items = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kasa Hareketleri"

    hf = PatternFill("solid", fgColor="1e40af")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    green_fill = PatternFill("solid", fgColor="dcfce7")
    red_fill   = PatternFill("solid", fgColor="fee2e2")

    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value="Laves Kimya - Kasa / Ciro Takibi")
    c.font = Font(bold=True, size=13, color="FFFFFF"); c.fill = hf
    c.alignment = Alignment(horizontal="center"); ws.row_dimensions[1].height = 22

    headers = ["Tarih", "Tür", "Ödeme Yöntemi", "Kategori", "Açıklama", "Tutar (₺)", ""]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal="center")

    PAY_TR = {"cash": "Nakit", "pos": "POS/Kart", "elden": "Elden", "other": "Diğer"}
    for i, f in enumerate(items, 3):
        row_fill = green_fill if f.flow_type == "income" else red_fill
        vals = [
            str(f.flow_date),
            "Gelir" if f.flow_type == "income" else "Gider",
            PAY_TR.get(f.pay_method, f.pay_method or "-"),
            f.category or "-",
            f.description or "-",
            f.amount,
            ""
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.fill = row_fill
            if col == 6: cell.number_format = '#,##0.00'

    # Özet
    total_income  = sum(f.amount for f in items if f.flow_type == "income")
    total_expense = sum(f.amount for f in items if f.flow_type == "expense")
    last = len(items) + 4
    ws.cell(row=last, column=5, value="Toplam Gelir:").font = Font(bold=True)
    ws.cell(row=last, column=6, value=round(total_income, 2)).font = Font(bold=True, color="166534")
    ws.cell(row=last+1, column=5, value="Toplam Gider:").font = Font(bold=True)
    ws.cell(row=last+1, column=6, value=round(total_expense, 2)).font = Font(bold=True, color="991b1b")
    ws.cell(row=last+2, column=5, value="Net:").font = Font(bold=True)
    ws.cell(row=last+2, column=6, value=round(total_income - total_expense, 2)).font = Font(bold=True)

    col_widths = [12, 10, 14, 16, 30, 14, 4]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kasa_hareketleri.xlsx"})
