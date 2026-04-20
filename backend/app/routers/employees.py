from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from datetime import date
from io import BytesIO
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.core.database import get_db
from app.core.deps import get_current_user, log_activity, check_permission
from app.models.employee import Employee, EmployeeLeave, SalaryPayment
from app.models.cashflow import CashFlow
from app.models.user import User

router = APIRouter(prefix="/employees", tags=["employees"])

# ── Pydantic Schemas ──────────────────────────────────────────────────────

class EmployeeCreate(BaseModel):
    first_name: str
    last_name: str
    tc_no: Optional[str] = None
    birth_date: Optional[date] = None
    gender: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    emergency_contact: Optional[str] = None
    emergency_phone: Optional[str] = None
    department: Optional[str] = None
    position: Optional[str] = None
    hire_date: Optional[date] = None
    salary: Optional[float] = 0
    iban: Optional[str] = None
    bank_name: Optional[str] = None
    annual_leave_days: Optional[int] = 14
    notes: Optional[str] = None

class EmployeeUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    tc_no: Optional[str] = None
    birth_date: Optional[date] = None
    gender: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    emergency_contact: Optional[str] = None
    emergency_phone: Optional[str] = None
    department: Optional[str] = None
    position: Optional[str] = None
    hire_date: Optional[date] = None
    termination_date: Optional[date] = None
    status: Optional[str] = None
    salary: Optional[float] = None
    iban: Optional[str] = None
    bank_name: Optional[str] = None
    annual_leave_days: Optional[int] = None
    notes: Optional[str] = None

class LeaveCreate(BaseModel):
    leave_type: str = "annual"
    start_date: date
    end_date: date
    days: int
    description: Optional[str] = None

class LeaveUpdate(BaseModel):
    status: Optional[str] = None
    description: Optional[str] = None

class SalaryCreate(BaseModel):
    period: str          # "2024-03"
    gross: float
    deductions: float = 0
    net: float
    paid_date: Optional[date] = None
    notes: Optional[str] = None

# ── Helpers ───────────────────────────────────────────────────────────────

def build_employee(e: Employee) -> dict:
    remaining_leave = e.annual_leave_days - e.used_leave_days
    return {
        "id": e.id,
        "first_name": e.first_name,
        "last_name": e.last_name,
        "full_name": f"{e.first_name} {e.last_name}",
        "tc_no": e.tc_no,
        "birth_date": e.birth_date,
        "gender": e.gender,
        "phone": e.phone,
        "email": e.email,
        "address": e.address,
        "emergency_contact": e.emergency_contact,
        "emergency_phone": e.emergency_phone,
        "department": e.department,
        "position": e.position,
        "hire_date": e.hire_date,
        "termination_date": e.termination_date,
        "status": e.status,
        "salary": e.salary,
        "iban": e.iban,
        "bank_name": e.bank_name,
        "annual_leave_days": e.annual_leave_days,
        "used_leave_days": e.used_leave_days,
        "remaining_leave": remaining_leave,
        "notes": e.notes,
        "created_at": e.created_at,
    }

# ── Employee CRUD ─────────────────────────────────────────────────────────

@router.get("")
def list_employees(
    skip: int = 0, limit: int = 50,
    search: Optional[str] = None,
    department: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(check_permission("employees", "view"))
):
    query = db.query(Employee)
    if search:
        query = query.filter(
            (Employee.first_name + " " + Employee.last_name).ilike(f"%{search}%") |
            Employee.phone.ilike(f"%{search}%") |
            Employee.department.ilike(f"%{search}%")
        )
    if department:
        query = query.filter(Employee.department == department)
    if status:
        query = query.filter(Employee.status == status)
    total = query.count()
    employees = query.order_by(Employee.first_name).offset(skip).limit(limit).all()
    return {"total": total, "items": [build_employee(e) for e in employees]}

@router.post("")
def create_employee(data: EmployeeCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    e = Employee(**data.model_dump())
    db.add(e); db.commit(); db.refresh(e)
    log_activity(db, current_user.id, "CREATE", "Employee", e.id, f"Personel eklendi: {e.first_name} {e.last_name}")
    return build_employee(e)

@router.get("/departments")
def list_departments(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    rows = db.query(Employee.department).filter(Employee.department != None).distinct().all()
    return sorted([r[0] for r in rows if r[0]])

@router.get("/{employee_id}")
def get_employee(employee_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    e = db.query(Employee).filter(Employee.id == employee_id).first()
    if not e: raise HTTPException(404, "Personel bulunamadı")
    result = build_employee(e)
    result["leaves"] = [build_leave(l) for l in e.leaves]
    result["salary_payments"] = [build_salary(s) for s in sorted(e.salary_payments, key=lambda x: x.period, reverse=True)]
    return result

@router.put("/{employee_id}")
def update_employee(employee_id: int, data: EmployeeUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    e = db.query(Employee).filter(Employee.id == employee_id).first()
    if not e: raise HTTPException(404, "Personel bulunamadı")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(e, k, v)
    db.commit(); db.refresh(e)
    log_activity(db, current_user.id, "UPDATE", "Employee", employee_id)
    return build_employee(e)

@router.delete("/{employee_id}")
def delete_employee(employee_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    e = db.query(Employee).filter(Employee.id == employee_id).first()
    if not e: raise HTTPException(404, "Personel bulunamadı")
    db.delete(e); db.commit()
    log_activity(db, current_user.id, "DELETE", "Employee", employee_id)
    return {"message": "Personel silindi"}

# ── İzin Yönetimi ─────────────────────────────────────────────────────────

def build_leave(l: EmployeeLeave) -> dict:
    TYPE_TR = {"annual":"Yıllık","sick":"Hastalık","excuse":"Mazeret","unpaid":"Ücretsiz","other":"Diğer"}
    STATUS_TR = {"pending":"Bekliyor","approved":"Onaylandı","rejected":"Reddedildi"}
    return {
        "id": l.id, "employee_id": l.employee_id,
        "leave_type": l.leave_type, "leave_type_label": TYPE_TR.get(l.leave_type, l.leave_type),
        "start_date": l.start_date, "end_date": l.end_date, "days": l.days,
        "status": l.status, "status_label": STATUS_TR.get(l.status, l.status),
        "description": l.description, "created_at": l.created_at,
    }

@router.post("/{employee_id}/leaves")
def create_leave(employee_id: int, data: LeaveCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    e = db.query(Employee).filter(Employee.id == employee_id).first()
    if not e: raise HTTPException(404, "Personel bulunamadı")
    leave = EmployeeLeave(employee_id=employee_id, **data.model_dump())
    db.add(leave); db.commit(); db.refresh(leave)
    return build_leave(leave)

@router.put("/{employee_id}/leaves/{leave_id}")
def update_leave(employee_id: int, leave_id: int, data: LeaveUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    leave = db.query(EmployeeLeave).filter(EmployeeLeave.id == leave_id, EmployeeLeave.employee_id == employee_id).first()
    if not leave: raise HTTPException(404, "İzin kaydı bulunamadı")
    old_status = leave.status
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(leave, k, v)
    # İzin onaylandıysa kullanılan izin günlerini güncelle
    if data.status == "approved" and old_status != "approved":
        e = db.query(Employee).filter(Employee.id == employee_id).first()
        if e and leave.leave_type == "annual":
            e.used_leave_days = (e.used_leave_days or 0) + leave.days
    # İzin reddedildiyse geri al
    elif data.status == "rejected" and old_status == "approved":
        e = db.query(Employee).filter(Employee.id == employee_id).first()
        if e and leave.leave_type == "annual":
            e.used_leave_days = max(0, (e.used_leave_days or 0) - leave.days)
    db.commit(); db.refresh(leave)
    return build_leave(leave)

@router.delete("/{employee_id}/leaves/{leave_id}")
def delete_leave(employee_id: int, leave_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    leave = db.query(EmployeeLeave).filter(EmployeeLeave.id == leave_id, EmployeeLeave.employee_id == employee_id).first()
    if not leave: raise HTTPException(404, "İzin kaydı bulunamadı")
    db.delete(leave); db.commit()
    return {"message": "Silindi"}

# ── Maaş Ödemeleri ────────────────────────────────────────────────────────

def build_salary(s: SalaryPayment) -> dict:
    return {
        "id": s.id, "employee_id": s.employee_id,
        "period": s.period, "gross": s.gross,
        "deductions": s.deductions, "net": s.net,
        "paid_date": s.paid_date, "notes": s.notes, "created_at": s.created_at,
    }

@router.post("/{employee_id}/salary")
def create_salary_payment(employee_id: int, data: SalaryCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    e = db.query(Employee).filter(Employee.id == employee_id).first()
    if not e: raise HTTPException(404, "Personel bulunamadı")
    sp = SalaryPayment(employee_id=employee_id, **data.model_dump())
    db.add(sp)
    # Kasaya otomatik gider olarak ekle
    cf = CashFlow(
        flow_date=data.paid_date or date.today(),
        flow_type="expense",
        amount=data.net,
        description=f"Maaş ödemesi: {e.first_name} {e.last_name} ({data.period})",
        category="Maaş",
    )
    db.add(cf)
    db.commit(); db.refresh(sp)
    log_activity(db, current_user.id, "SALARY", "Employee", employee_id,
                 f"{e.first_name} {e.last_name} — {data.period} maaşı ödendi: ₺{data.net}")
    return build_salary(sp)

@router.delete("/{employee_id}/salary/{salary_id}")
def delete_salary(employee_id: int, salary_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    sp = db.query(SalaryPayment).filter(SalaryPayment.id == salary_id, SalaryPayment.employee_id == employee_id).first()
    if not sp: raise HTTPException(404, "Maaş kaydı bulunamadı")
    db.delete(sp); db.commit()
    return {"message": "Silindi"}

# ── Excel Export ──────────────────────────────────────────────────────────

@router.get("/export/list")
def export_employees(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    employees = db.query(Employee).order_by(Employee.first_name).all()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Personel"
    hf = PatternFill("solid", fgColor="1e40af")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value="Laves Kimya - Personel Listesi")
    c.font = Font(bold=True, size=13, color="FFFFFF"); c.fill = hf
    c.alignment = Alignment(horizontal="center"); ws.row_dimensions[1].height = 22
    headers = ["Ad Soyad","Departman","Pozisyon","Telefon","E-posta","İşe Giriş","Durum","Brüt Maaş","IBAN","Banka","Yıllık İzin","Kalan İzin"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal="center")
    STATUS_TR = {"active":"Aktif","inactive":"Pasif","on_leave":"İzinde"}
    alt = PatternFill("solid", fgColor="f0f4ff")
    for i, e in enumerate(employees, 3):
        row_fill = alt if i % 2 == 0 else None
        vals = [
            f"{e.first_name} {e.last_name}", e.department or "-", e.position or "-",
            e.phone or "-", e.email or "-",
            str(e.hire_date) if e.hire_date else "-",
            STATUS_TR.get(e.status, e.status),
            e.salary, e.iban or "-", e.bank_name or "-",
            e.annual_leave_days, e.annual_leave_days - e.used_leave_days
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            if row_fill: cell.fill = row_fill
            if col == 8: cell.number_format = '#,##0.00'
    col_widths = [22,16,18,14,24,12,10,14,26,16,12,12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=personel_listesi.xlsx"})
