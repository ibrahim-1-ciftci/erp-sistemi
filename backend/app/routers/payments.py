from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from datetime import date
from io import BytesIO
from pydantic import BaseModel
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.database import get_db
from app.core.deps import get_current_user, check_permission, check_permission, log_activity
from app.models.payment import Payment, PaymentStatus
from app.models.order import Order
from app.models.user import User

router = APIRouter(prefix="/payments", tags=["payments"])

class PaymentItem(BaseModel):
    product_name: str
    quantity: float
    unit: str = "adet"
    unit_price: float

class PaymentCreate(BaseModel):
    order_id: Optional[int] = None
    customer_name_override: Optional[str] = None
    order_date: Optional[date] = None
    total_amount: float
    due_date: date
    items: Optional[List[PaymentItem]] = []
    notes: Optional[str] = None

class PaymentUpdate(BaseModel):
    paid_amount: Optional[float] = None
    paid_date: Optional[date] = None
    due_date: Optional[date] = None
    order_date: Optional[date] = None
    total_amount: Optional[float] = None
    notes: Optional[str] = None
    items_json: Optional[str] = None

    class Config:
        # None değerleri de set edilebilsin (arşivden çıkar için paid_date=None gerekli)
        # exclude_unset kullanıyoruz, exclude_none değil
        pass

def compute_status(p: Payment) -> str:
    paid = p.paid_amount or 0
    if paid >= p.total_amount:
        return PaymentStatus.paid
    if paid > 0:
        return PaymentStatus.partial
    if p.due_date < date.today():
        return PaymentStatus.overdue
    return PaymentStatus.pending

def build_out(p: Payment) -> dict:
    status = compute_status(p)
    remaining = p.total_amount - (p.paid_amount or 0)
    overdue_days = (date.today() - p.due_date).days if p.due_date < date.today() and status != PaymentStatus.paid else 0
    items = []
    if p.items_json:
        try:
            items = json.loads(p.items_json)
        except Exception:
            items = []
    return {
        "id": p.id,
        "order_id": p.order_id,
        "customer_name": p.customer_name,
        "order_date": p.order_date,
        "items": items,
        "total_amount": p.total_amount,
        "paid_amount": p.paid_amount or 0,
        "remaining": remaining,
        "due_date": p.due_date,
        "paid_date": p.paid_date,
        "status": status,
        "notes": p.notes,
        "created_at": p.created_at,
        "overdue_days": overdue_days,
    }

@router.get("")
def list_payments(status: Optional[str] = None, db: Session = Depends(get_db), current_user: User = Depends(check_permission("payments", "view"))):
    payments = db.query(Payment).order_by(Payment.due_date).all()
    result = [build_out(p) for p in payments]
    if status:
        result = [r for r in result if r["status"] == status]
    total_pending = sum(r["remaining"] for r in result if r["status"] != PaymentStatus.paid)
    overdue_count = sum(1 for r in result if r["status"] == PaymentStatus.overdue)
    return {"items": result, "total_pending": total_pending, "overdue_count": overdue_count}

@router.get("/calendar")
def calendar_view(year: int, month: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    payments = db.query(Payment).filter(
        Payment.due_date >= date(year, month, 1),
        Payment.due_date <= date(year, month, last_day)
    ).all()
    by_day = {}
    for p in payments:
        day = p.due_date.day
        if day not in by_day:
            by_day[day] = []
        by_day[day].append(build_out(p))
    return {"year": year, "month": month, "by_day": by_day}

@router.post("")
def create_payment(data: PaymentCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if data.order_id:
        order = db.query(Order).filter(Order.id == data.order_id).first()
        if not order:
            raise HTTPException(status_code=404, detail="Sipariş bulunamadı")
        customer_name = data.customer_name_override or order.customer_name
    elif data.customer_name_override:
        customer_name = data.customer_name_override
    else:
        raise HTTPException(status_code=400, detail="Sipariş veya müşteri adı gerekli")

    items_json = json.dumps([i.model_dump() for i in (data.items or [])], ensure_ascii=False)
    payment = Payment(
        order_id=data.order_id,
        customer_name=customer_name,
        order_date=data.order_date,
        total_amount=data.total_amount,
        due_date=data.due_date,
        items_json=items_json,
        notes=data.notes
    )
    db.add(payment)
    db.commit()
    db.refresh(payment)
    log_activity(db, current_user.id, "CREATE", "Payment", payment.id, f"Vade oluşturuldu: {customer_name}")
    return build_out(payment)

@router.put("/{payment_id}")
def update_payment(payment_id: int, data: PaymentUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    p = db.query(Payment).filter(Payment.id == payment_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")

    old_paid = p.paid_amount or 0
    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(p, key, value)

    # Tam ödeme yapıldıysa ve paid_date yoksa bugünü ata
    if data.paid_amount is not None and data.paid_amount >= p.total_amount and not p.paid_date:
        p.paid_date = date.today()

    # Yeni ödeme tutarı varsa farkı kasaya gelir olarak ekle
    if data.paid_amount is not None:
        new_paid = data.paid_amount or 0
        diff = new_paid - old_paid
        if diff > 0:
            from app.models.cashflow import CashFlow
            cf = CashFlow(
                flow_date=data.paid_date or date.today(),
                flow_type="income",
                amount=diff,
                description=f"Vade tahsilatı: {p.customer_name}",
                category="Vade Tahsilatı",
            )
            db.add(cf)

    db.commit()
    db.refresh(p)
    log_activity(db, current_user.id, "UPDATE", "Payment", payment_id)
    return build_out(p)

@router.delete("/{payment_id}")
def delete_payment(payment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    p = db.query(Payment).filter(Payment.id == payment_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
    db.delete(p)
    db.commit()
    return {"message": "Silindi"}

@router.post("/invoice-pdf")
def payment_invoice_pdf(
    payment_ids: list[int],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    import os, time
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from app.routers.settings import get_setting
    from datetime import datetime as dt

    if not payment_ids:
        raise HTTPException(400, "En az bir kayit secin")
    payments_list = db.query(Payment).filter(Payment.id.in_(payment_ids)).all()
    if not payments_list:
        raise HTTPException(404, "Kayit bulunamadi")

    company_name = get_setting(db, "company_name") or "Laves Kimya"
    company_sub  = get_setting(db, "company_sub")  or "Uretim ve Isletme Yonetim Sistemi"
    kdv_rate     = float(get_setting(db, "kdv_rate") or "20")

    FN, FB = "Helvetica", "Helvetica-Bold"
    for reg, bold in [
        ("C:/Windows/Fonts/arial.ttf",   "C:/Windows/Fonts/arialbd.ttf"),
        ("C:/Windows/Fonts/calibri.ttf", "C:/Windows/Fonts/calibrib.ttf"),
        ("C:/Windows/Fonts/tahoma.ttf",  "C:/Windows/Fonts/tahomabd.ttf"),
    ]:
        if os.path.exists(reg):
            bname = os.path.splitext(os.path.basename(reg))[0].capitalize()
            try:
                registered = pdfmetrics.getRegisteredFontNames()
                if bname not in registered:
                    pdfmetrics.registerFont(TTFont(bname, reg))
                bname_bold = bname + "-Bold"
                if os.path.exists(bold) and bname_bold not in registered:
                    pdfmetrics.registerFont(TTFont(bname_bold, bold))
                FN = bname
                FB = bname_bold if os.path.exists(bold) else bname
                break
            except Exception:
                continue

    uid = str(int(time.time() * 1000))
    BLUE  = colors.HexColor('#1e40af')
    LIGHT = colors.HexColor('#f8fafc')
    BORD  = colors.HexColor('#e2e8f0')
    GREEN = colors.HexColor('#16a34a')
    RED   = colors.HexColor('#dc2626')

    def ps(name, fn=None, fb=False, size=9, color=None, align=None):
        kw = {"fontName": fb and FB or (fn or FN), "fontSize": size}
        if color: kw["textColor"] = color
        if align: kw["alignment"] = align
        return ParagraphStyle(f"pi_{name}_{uid}", **kw)

    S_title = ps("title", fb=True, size=18, color=BLUE)
    S_sub   = ps("sub",   size=8,  color=colors.grey)
    S_right = ps("right", size=10, align=TA_RIGHT)
    S_bold  = ps("bold",  fb=True, size=10)
    S_norm  = ps("norm",  size=9)
    S_ra    = ps("ra",    size=9,  align=TA_RIGHT)
    S_ra_g  = ps("rag",   size=9,  align=TA_RIGHT, color=GREEN)
    S_ra_r  = ps("rar",   fb=True, size=10, align=TA_RIGHT, color=RED)
    S_gt    = ps("gt",    fb=True, size=10)
    S_gtr   = ps("gtr",   fb=True, size=10, align=TA_RIGHT)
    S_hdr   = ps("hdr",   fb=True, size=9,  color=colors.white)
    S_hdrr  = ps("hdrr",  fb=True, size=9,  color=colors.white, align=TA_RIGHT)
    S_foot  = ps("foot",  size=7,  color=colors.grey, align=TA_CENTER)
    S_small = ps("small", size=8,  color=colors.grey)

    STATUS_TR = {"pending":"Bekliyor","overdue":"Gecikmiş","partial":"Kısmi Ödeme","paid":"Ödendi"}

    story = []
    for idx, p in enumerate(payments_list):
        if idx > 0:
            story.append(PageBreak())

        items = []
        if p.items_json:
            try:
                items = json.loads(p.items_json)
            except Exception:
                items = []

        hdr_tbl = Table([
            [Paragraph(company_name, S_title), Paragraph(f'VADE FATURASI  #{p.id:05d}', S_right)],
            [Paragraph(company_sub,  S_sub),   Paragraph(f'Tarih: {dt.now().strftime("%d.%m.%Y")}', S_right)],
        ], colWidths=[10*cm, 7*cm])
        hdr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP')]))
        story.append(hdr_tbl)
        story.append(Spacer(1, 0.3*cm))

        ln = Table([['']], colWidths=[17*cm])
        ln.setStyle(TableStyle([('LINEABOVE',(0,0),(-1,0),1.5,BLUE)]))
        story.append(ln)
        story.append(Spacer(1, 0.4*cm))

        status_val = build_out(p)['status']
        status_label = STATUS_TR.get(str(status_val), str(status_val))
        info_tbl = Table([
            [Paragraph(f'Musteri: <b>{p.customer_name}</b>', S_norm),
             Paragraph(f'Durum: <b>{status_label}</b>', S_norm)],
            [Paragraph(f'Siparis Tarihi: {p.order_date or "-"}', S_norm),
             Paragraph(f'Vade Tarihi: {p.due_date}', S_norm)],
        ], colWidths=[9*cm, 8*cm])
        info_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1),LIGHT),
            ('BOX',(0,0),(-1,-1),0.5,BORD),
            ('LEFTPADDING',(0,0),(-1,-1),8),
            ('TOPPADDING',(0,0),(-1,-1),5),
            ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]))
        story.append(info_tbl)
        story.append(Spacer(1, 0.5*cm))

        tdata = [[
            Paragraph('#',           S_hdr),
            Paragraph('Urun Adi',    S_hdr),
            Paragraph('Miktar',      S_hdrr),
            Paragraph('Birim',       S_hdrr),
            Paragraph('Birim Fiyat', S_hdrr),
            Paragraph('Toplam',      S_hdrr),
        ]]
        subtotal = 0.0
        if items:
            for i, item in enumerate(items, 1):
                qty   = float(item.get('quantity', 1))
                price = float(item.get('unit_price', 0))
                line  = qty * price
                subtotal += line
                tdata.append([
                    str(i),
                    Paragraph(str(item.get('product_name','')), S_norm),
                    Paragraph(f"{qty:g}",          S_ra),
                    Paragraph(str(item.get('unit','adet')), S_ra),
                    Paragraph(f"{price:,.2f} TL",  S_ra),
                    Paragraph(f"{line:,.2f} TL",   S_ra),
                ])
        else:
            subtotal = float(p.total_amount)
            tdata.append(['1', Paragraph('(Urun detayi girilmemis)', S_norm),
                          '-', '-', '-', Paragraph(f"{subtotal:,.2f} TL", S_ra)])

        n_items = len(tdata)
        kdv_amt     = subtotal * (kdv_rate / 100)
        total_w_kdv = subtotal + kdv_amt
        paid_amt    = float(p.paid_amount or 0)
        remaining   = float(p.total_amount) - paid_amt

        tdata += [
            ['','','','', Paragraph('Ara Toplam:',          S_bold), Paragraph(f'{subtotal:,.2f} TL',    S_ra)],
            ['','','','', Paragraph(f'KDV (%{int(kdv_rate)}):', S_bold), Paragraph(f'{kdv_amt:,.2f} TL', S_ra)],
            ['','','','', Paragraph('GENEL TOPLAM:',        S_gt),   Paragraph(f'{total_w_kdv:,.2f} TL', S_gtr)],
            ['','','','', Paragraph('Odenen:',              S_norm), Paragraph(f'{paid_amt:,.2f} TL',    S_ra_g)],
            ['','','','', Paragraph('Kalan:',               S_bold), Paragraph(f'{remaining:,.2f} TL',   S_ra_r)],
        ]
        n_total = len(tdata)

        prod_tbl = Table(tdata, colWidths=[0.8*cm, 6.5*cm, 1.8*cm, 1.8*cm, 3.2*cm, 3*cm])
        prod_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),  (-1,0),           BLUE),
            ('FONTNAME',      (0,0),  (-1,-1),           FN),
            ('FONTSIZE',      (0,1),  (-1,-1),           9),
            ('ROWBACKGROUNDS',(0,1),  (-1,n_items-1),    [colors.white, LIGHT]),
            ('GRID',          (0,0),  (-1,n_items-1),    0.4, BORD),
            ('LINEABOVE',     (0,n_items),(-1,n_items),  1, BORD),
            ('BACKGROUND',    (4,n_total-3),(-1,n_total-3), colors.HexColor('#eff6ff')),
            ('TOPPADDING',    (0,0),  (-1,-1),           5),
            ('BOTTOMPADDING', (0,0),  (-1,-1),           5),
            ('LEFTPADDING',   (0,0),  (-1,-1),           5),
        ]))
        story.append(prod_tbl)
        story.append(Spacer(1, 0.6*cm))

        if p.notes:
            story.append(Paragraph(f'Not: {p.notes}', S_small))
            story.append(Spacer(1, 0.3*cm))

        story.append(Spacer(1, 0.4*cm))
        story.append(Paragraph(f'Bu belge {company_name} tarafindan otomatik olusturulmustur.', S_foot))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    doc.build(story)
    buf.seek(0)
    out_name = f"vade_fatura_{dt.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={out_name}"})
@router.get("/export")
def export_payments(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    payments = db.query(Payment).order_by(Payment.due_date).all()
    result = [build_out(p) for p in payments]
    if status:
        result = [r for r in result if r["status"] == status]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vade Takibi"

    # Şirket başlığı
    ws.merge_cells("A1:H1")
    title_c = ws.cell(row=1, column=1, value="Laves Kimya - Vade Takibi")
    title_c.font = Font(bold=True, size=14, color="FFFFFF")
    title_c.fill = PatternFill("solid", fgColor="1e40af")
    title_c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Stiller
    header_fill = PatternFill("solid", fgColor="1e40af")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    sub_fill = PatternFill("solid", fgColor="dbeafe")
    sub_font = Font(bold=True, size=9, color="1e3a8a")
    thin = Side(style="thin", color="d1d5db")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    STATUS_TR = {"pending": "Bekliyor", "overdue": "Gecikmiş", "partial": "Kısmi Ödeme", "paid": "Ödendi"}

    row = 3
    for p in result:
        # Müşteri başlık satırı
        ws.merge_cells(f"A{row}:H{row}")
        title_cell = ws.cell(row=row, column=1,
            value=f"Müşteri: {p['customer_name']}  |  Sipariş Tarihi: {p['order_date'] or '-'}  |  Vade: {p['due_date']}  |  Durum: {STATUS_TR.get(p['status'], p['status'])}")
        title_cell.fill = header_fill
        title_cell.font = header_font
        title_cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 20
        row += 1

        # Ürün başlıkları
        headers = ["Ürün Adı", "Miktar", "Birim", "Birim Fiyat (₺)", "Toplam (₺)", "", "", ""]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = sub_fill
            c.font = sub_font
            c.border = border
            c.alignment = center
        row += 1

        # Ürün satırları
        if p["items"]:
            for item in p["items"]:
                total = item["quantity"] * item["unit_price"]
                vals = [item["product_name"], item["quantity"], item["unit"], item["unit_price"], total, "", "", ""]
                for col, v in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.border = border
                    if col in (4, 5):
                        c.number_format = '#,##0.00 ₺'
                row += 1
        else:
            ws.cell(row=row, column=1, value="(ürün detayı girilmemiş)").font = Font(italic=True, color="9ca3af")
            row += 1

        # Özet satırı
        summary_vals = ["", "", "", "Toplam:", f"₺{p['total_amount']:,.2f}",
                        f"Ödenen: ₺{p['paid_amount']:,.2f}", f"Kalan: ₺{p['remaining']:,.2f}", ""]
        for col, v in enumerate(summary_vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = Font(bold=True, size=9)
            if col in (4, 5, 6, 7):
                c.fill = PatternFill("solid", fgColor="f0fdf4")
        row += 1

        # Notlar
        if p["notes"]:
            ws.cell(row=row, column=1, value=f"Not: {p['notes']}").font = Font(italic=True, size=8, color="6b7280")
            row += 1

        # Boş ayırıcı satır
        row += 1

    col_widths = [30, 10, 10, 16, 16, 18, 18, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vade_takibi.xlsx"}
    )
